"""Freelancer Finance Tracker — costruzione del workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, DataBarRule

NAVY   = "1F3864"
BLUE   = "2E5C8A"
LIGHT  = "D9E2F3"
GREY   = "F2F2F2"
GREEN  = "375623"
RED    = "C00000"
INPUT  = "FFF2CC"      # celle da compilare
WHITE  = "FFFFFF"

F = "Arial"
CUR = '"$"#,##0.00;("$"#,##0.00);"$"0.00'
CUR0 = '"$"#,##0;("$"#,##0);"$"0'
PCT = '0.0%'

thin = Side(style="thin", color="BFBFBF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def titolo(ws, testo, riga=1, span=6):
    ws.merge_cells(start_row=riga, start_column=1, end_row=riga, end_column=span)
    c = ws.cell(riga, 1, testo)
    c.font = Font(F, size=16, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[riga].height = 30

def intestazioni(ws, riga, valori, larghezze):
    for i, (v, w) in enumerate(zip(valori, larghezze), start=1):
        c = ws.cell(riga, i, v)
        c.font = Font(F, size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = box
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[riga].height = 28

# =====================================================================
# 1. START HERE
# =====================================================================
ws = wb.active
ws.title = "Start Here"
ws.sheet_view.showGridLines = False
titolo(ws, "  Freelancer Finance Tracker", span=8)

righe = [
    ("", ""),
    ("HOW THIS WORKBOOK WORKS", ""),
    ("", "You only ever type in three sheets. Everything else calculates itself."),
    ("", ""),
    ("1.  Settings", "Set your tax rate and edit the category lists. Do this once."),
    ("2.  Income", "Log every invoice. One row per invoice."),
    ("3.  Expenses", "Log every business expense. One row per expense."),
    ("", ""),
    ("Dashboard", "Reads automatically. Never type here."),
    ("Monthly Summary", "Reads automatically. Never type here."),
    ("Clients", "Reads automatically. Shows which clients are actually worth it."),
    ("", ""),
    ("COLOUR KEY", ""),
    ("", "Yellow cells are for you to fill in."),
    ("", "White cells contain formulas. Typing over one breaks it."),
    ("", ""),
    ("BEFORE YOU START", ""),
    ("", "Each log sheet has one example row so you can see the expected format."),
    ("", "Delete it once you have entered your own data."),
    ("", ""),
    ("A NOTE ON TAX", ""),
    ("", "The tax set-aside is a planning estimate based on the rate you enter."),
    ("", "It is not tax advice and does not replace an accountant."),
]
r = 3
for a, b in righe:
    ws.cell(r, 1, a).font = Font(F, size=11, bold=bool(a and a.isupper()),
                                 color=NAVY if a else "000000")
    ws.cell(r, 2, b).font = Font(F, size=10)
    r += 1
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 72
c = ws.cell(16, 2); c.fill = PatternFill("solid", fgColor=INPUT); c.border = box

# =====================================================================
# 2. SETTINGS
# =====================================================================
st = wb.create_sheet("Settings")
st.sheet_view.showGridLines = False
titolo(st, "  Settings", span=4)
st["A3"] = "Set these once"; st["A3"].font = Font(F, size=12, bold=True, color=NAVY)

st["A5"] = "Tax set-aside rate"
st["B5"] = 0.25
st["C5"] = "Percentage of profit to reserve for tax. Ask your accountant."
st["A6"] = "Business name"
st["B6"] = "Your Business"
st["A7"] = "Currency symbol shown"
st["B7"] = "$"
st["C7"] = "Cosmetic only. Change cell formats to switch currency fully."

for rr in (5, 6, 7):
    st.cell(rr, 1).font = Font(F, size=10, bold=True)
    c = st.cell(rr, 2); c.fill = PatternFill("solid", fgColor=INPUT)
    c.font = Font(F, size=10, color="0000FF"); c.border = box
    c.alignment = Alignment(horizontal="center")
    st.cell(rr, 3).font = Font(F, size=9, italic=True, color="808080")
st["B5"].number_format = PCT

st["A10"] = "Income categories"; st["A10"].font = Font(F, size=11, bold=True, color=NAVY)
st["C10"] = "Expense categories"; st["C10"].font = Font(F, size=11, bold=True, color=NAVY)
inc_cat = ["Client work","Retainer","Consulting","Royalties","Teaching","Affiliate","Other"]
exp_cat = ["Software","Hardware","Subcontractors","Marketing","Travel","Office",
           "Professional fees","Training","Insurance","Bank fees","Other"]
for i, v in enumerate(inc_cat):
    c = st.cell(11+i, 1, v); c.font = Font(F, size=10); c.border = box
    c.fill = PatternFill("solid", fgColor=INPUT)
for i, v in enumerate(exp_cat):
    c = st.cell(11+i, 3, v); c.font = Font(F, size=10); c.border = box
    c.fill = PatternFill("solid", fgColor=INPUT)
for col, w in zip("ABCD", [22, 16, 22, 50]): st.column_dimensions[col].width = w

# =====================================================================
# 3. INCOME
# =====================================================================
N = 300
inc = wb.create_sheet("Income")
inc.sheet_view.showGridLines = False
titolo(inc, "  Income  —  one row per invoice", span=7)
inc["A2"] = "Yellow = type here.  Leave 'Date paid' empty until the invoice is actually paid."
inc["A2"].font = Font(F, size=9, italic=True, color="808080")
intestazioni(inc, 3, ["Date issued","Client","Description","Category",
                      "Amount","Date paid","Status"], [14,22,34,18,14,14,14])
inc.append_ex = None
esempio = ["2026-01-15","Acme Ltd","Website redesign — phase 1","Client work",1500,"2026-02-02"]
for i, v in enumerate(esempio, start=1):
    c = inc.cell(4, i, v); c.font = Font(F, size=10, italic=True, color="808080"); c.border = box
inc.cell(4,5).number_format = CUR
for r in range(5, 4+N):
    for cc in range(1, 7):
        c = inc.cell(r, cc); c.border = box
        c.fill = PatternFill("solid", fgColor=INPUT); c.font = Font(F, size=10)
    inc.cell(r,5).number_format = CUR
for r in range(4, 4+N):
    inc.cell(r,7).value = f'=IF(E{r}="","",IF(F{r}="","Unpaid","Paid"))'
    inc.cell(r,7).font = Font(F, size=10, bold=True)
    inc.cell(r,7).alignment = Alignment(horizontal="center")
    inc.cell(r,7).border = box
inc.freeze_panes = "A4"
dv = DataValidation(type="list", formula1="=Settings!$A$11:$A$17", allow_blank=True)
inc.add_data_validation(dv); dv.add(f"D4:D{3+N}")
inc.conditional_formatting.add(f"G4:G{3+N}",
    CellIsRule(operator="equal", formula=['"Paid"'], font=Font(F, size=10, bold=True, color=GREEN)))
inc.conditional_formatting.add(f"G4:G{3+N}",
    CellIsRule(operator="equal", formula=['"Unpaid"'], font=Font(F, size=10, bold=True, color=RED)))

# =====================================================================
# 4. EXPENSES
# =====================================================================
exp = wb.create_sheet("Expenses")
exp.sheet_view.showGridLines = False
titolo(exp, "  Expenses  —  one row per expense", span=6)
exp["A2"] = "Yellow = type here.  Mark 'Deductible' as Yes or No."
exp["A2"].font = Font(F, size=9, italic=True, color="808080")
intestazioni(exp, 3, ["Date","Supplier","Description","Category","Amount","Deductible"],
             [14,22,34,20,14,13])
esempio = ["2026-01-08","Adobe","Creative Cloud subscription","Software",59.99,"Yes"]
for i, v in enumerate(esempio, start=1):
    c = exp.cell(4, i, v); c.font = Font(F, size=10, italic=True, color="808080"); c.border = box
exp.cell(4,5).number_format = CUR
for r in range(5, 4+N):
    for cc in range(1, 7):
        c = exp.cell(r, cc); c.border = box
        c.fill = PatternFill("solid", fgColor=INPUT); c.font = Font(F, size=10)
    exp.cell(r,5).number_format = CUR
exp.freeze_panes = "A4"
dv2 = DataValidation(type="list", formula1="=Settings!$C$11:$C$21", allow_blank=True)
exp.add_data_validation(dv2); dv2.add(f"D4:D{3+N}")
dv3 = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
exp.add_data_validation(dv3); dv3.add(f"F4:F{3+N}")

wb.save("Freelancer-Finance-Tracker.xlsx")
print("struttura base creata")
