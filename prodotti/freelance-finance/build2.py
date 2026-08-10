"""Aggiunge Dashboard, Monthly Summary e Clients — tutto a formule."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, CellIsRule
from openpyxl.chart import BarChart, LineChart, Reference

NAVY="1F3864"; BLUE="2E5C8A"; LIGHT="D9E2F3"; GREEN="375623"; RED="C00000"; WHITE="FFFFFF"
F="Arial"; CUR='"$"#,##0.00;("$"#,##0.00);"$"0.00'; CUR0='"$"#,##0;("$"#,##0);"$"0'; PCT='0.0%'
thin=Side(style="thin",color="BFBFBF"); box=Border(left=thin,right=thin,top=thin,bottom=thin)
N=300; R=f"4:{3+N}"

wb=load_workbook("Freelancer-Finance-Tracker.xlsx")

def titolo(ws,t,span=6):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=span)
    c=ws.cell(1,1,t); c.font=Font(F,size=16,bold=True,color=WHITE)
    c.fill=PatternFill("solid",fgColor=NAVY)
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[1].height=30

def hdr(ws,riga,vals,ws_):
    for i,(v,w) in enumerate(zip(vals,ws_),start=1):
        c=ws.cell(riga,i,v); c.font=Font(F,size=10,bold=True,color=WHITE)
        c.fill=PatternFill("solid",fgColor=BLUE); c.border=box
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[riga].height=26

# ---------------------------------------------------------------- MONTHLY
ms=wb.create_sheet("Monthly Summary"); ms.sheet_view.showGridLines=False
titolo(ms,"  Monthly Summary  —  calculated automatically",span=7)
ms["A2"]="Every figure here is a formula. Do not type in this sheet."
ms["A2"].font=Font(F,size=9,italic=True,color="808080")
hdr(ms,3,["Month","Income invoiced","Income received","Expenses","Profit",
          "Tax set-aside","Net after tax"],[14,17,17,15,15,15,16])
mesi=["2026-01","2026-02","2026-03","2026-04","2026-05","2026-06",
      "2026-07","2026-08","2026-09","2026-10","2026-11","2026-12"]
for i,m in enumerate(mesi):
    r=4+i
    ms.cell(r,1,m).font=Font(F,size=10,bold=True)
    ms.cell(r,1).alignment=Alignment(horizontal="center")
    # SUMPRODUCT su testo del mese: robusto e compatibile
    ms.cell(r,2,f'=SUMPRODUCT((LEFT(Income!$A${R.split(":")[0]}:$A${3+N},7)=$A{r})*Income!$E$4:$E${3+N})')
    ms.cell(r,3,f'=SUMPRODUCT((LEFT(Income!$F$4:$F${3+N},7)=$A{r})*Income!$E$4:$E${3+N})')
    ms.cell(r,4,f'=SUMPRODUCT((LEFT(Expenses!$A$4:$A${3+N},7)=$A{r})*Expenses!$E$4:$E${3+N})')
    ms.cell(r,5,f'=B{r}-D{r}')
    ms.cell(r,6,f'=IF(E{r}>0,E{r}*Settings!$B$5,0)')
    ms.cell(r,7,f'=E{r}-F{r}')
    for cc in range(2,8):
        c=ms.cell(r,cc); c.number_format=CUR0; c.border=box; c.font=Font(F,size=10)
    ms.cell(r,1).border=box
tr=4+len(mesi)
ms.cell(tr,1,"TOTAL").font=Font(F,size=11,bold=True,color=WHITE)
ms.cell(tr,1).fill=PatternFill("solid",fgColor=NAVY)
ms.cell(tr,1).alignment=Alignment(horizontal="center")
for cc in range(2,8):
    col=get_column_letter(cc)
    c=ms.cell(tr,cc,f'=SUM({col}4:{col}{tr-1})')
    c.font=Font(F,size=11,bold=True,color=WHITE); c.fill=PatternFill("solid",fgColor=NAVY)
    c.number_format=CUR0; c.border=box
ms.conditional_formatting.add(f"E4:E{tr-1}",
    CellIsRule(operator="lessThan",formula=["0"],font=Font(F,size=10,color=RED,bold=True)))
ms.freeze_panes="A4"

ch=LineChart(); ch.title="Profit by month"; ch.height=8; ch.width=20
ch.y_axis.numFmt=CUR0
d=Reference(ms,min_col=5,min_row=3,max_row=tr-1)
cats=Reference(ms,min_col=1,min_row=4,max_row=tr-1)
ch.add_data(d,titles_from_data=True); ch.set_categories(cats)
ms.add_chart(ch,"I3")

# ---------------------------------------------------------------- CLIENTS
cl=wb.create_sheet("Clients"); cl.sheet_view.showGridLines=False
titolo(cl,"  Clients  —  who is actually worth your time",span=6)
cl["A2"]="Type a client name in column A exactly as written in the Income sheet. The rest calculates."
cl["A2"].font=Font(F,size=9,italic=True,color="808080")
hdr(cl,3,["Client","Invoiced","Received","Outstanding","Invoices","% of total income"],
    [26,16,16,16,11,17])
for i in range(30):
    r=4+i
    c=cl.cell(r,1); c.fill=PatternFill("solid",fgColor="FFF2CC"); c.border=box; c.font=Font(F,size=10)
    cl.cell(r,2,f'=IF($A{r}="","",SUMIF(Income!$B$4:$B${3+N},$A{r},Income!$E$4:$E${3+N}))')
    cl.cell(r,3,f'=IF($A{r}="","",SUMPRODUCT((Income!$B$4:$B${3+N}=$A{r})*(Income!$F$4:$F${3+N}<>"")*Income!$E$4:$E${3+N}))')
    cl.cell(r,4,f'=IF($A{r}="","",B{r}-C{r})')
    cl.cell(r,5,f'=IF($A{r}="","",COUNTIF(Income!$B$4:$B${3+N},$A{r}))')
    cl.cell(r,6,f'=IF(OR($A{r}="",SUM(Income!$E$4:$E${3+N})=0),"",B{r}/SUM(Income!$E$4:$E${3+N}))')
    for cc in range(2,5):
        c=cl.cell(r,cc); c.number_format=CUR0; c.border=box; c.font=Font(F,size=10)
    cl.cell(r,5).border=box; cl.cell(r,5).alignment=Alignment(horizontal="center")
    cl.cell(r,5).font=Font(F,size=10)
    cl.cell(r,6).number_format=PCT; cl.cell(r,6).border=box; cl.cell(r,6).font=Font(F,size=10)
cl.conditional_formatting.add("B4:B33",
    DataBarRule(start_type="min",end_type="max",color=BLUE))
cl.conditional_formatting.add("D4:D33",
    CellIsRule(operator="greaterThan",formula=["0"],font=Font(F,size=10,color=RED,bold=True)))
cl.freeze_panes="A4"

# ---------------------------------------------------------------- DASHBOARD
db=wb.create_sheet("Dashboard",1); db.sheet_view.showGridLines=False
titolo(db,"  Dashboard",span=8)
db["A2"]="Everything on this sheet is calculated. Enter your data in Income and Expenses."
db["A2"].font=Font(F,size=9,italic=True,color="808080")

kpi=[("Total invoiced",  f'=SUM(Income!E4:E{3+N})'),
     ("Total received",   f'=SUMPRODUCT((Income!$F$4:$F${3+N}<>"")*Income!$E$4:$E${3+N})'),
     ("Outstanding",      '=B5-B6'),
     ("Total expenses",   f'=SUM(Expenses!E4:E{3+N})'),
     ("Profit",           '=B5-B8'),
     ("Tax set-aside",    '=IF(B9>0,B9*Settings!B5,0)'),
     ("Net after tax",    '=B9-B10'),
     ("Profit margin",    '=IFERROR(B9/B5,0)')]
r=5
for nome,form in kpi:
    c=db.cell(r,1,nome); c.font=Font(F,size=11,bold=True,color=NAVY)
    c.fill=PatternFill("solid",fgColor=LIGHT); c.border=box
    c.alignment=Alignment(vertical="center",indent=1)
    v=db.cell(r,2,form); v.font=Font(F,size=12,bold=True); v.border=box
    v.number_format=PCT if nome=="Profit margin" else CUR
    v.alignment=Alignment(horizontal="right",indent=1)
    db.row_dimensions[r].height=22
    r+=1
db.column_dimensions["A"].width=22; db.column_dimensions["B"].width=18
db.cell(9,2).font=Font(F,size=12,bold=True,color=GREEN)
db.conditional_formatting.add("B9:B9",
    CellIsRule(operator="lessThan",formula=["0"],font=Font(F,size=12,bold=True,color=RED)))

db["D5"]="Unpaid invoices"; db["D5"].font=Font(F,size=11,bold=True,color=NAVY)
db["E5"]=f'=COUNTIF(Income!G4:G{3+N},"Unpaid")'
db["D6"]="Amount outstanding"; db["D6"].font=Font(F,size=11,bold=True,color=NAVY)
db["E6"]='=B7'
db["D7"]="Active clients"; db["D7"].font=Font(F,size=11,bold=True,color=NAVY)
db["E7"]=f'=SUMPRODUCT((Income!$B$4:$B${3+N}<>"")/COUNTIF(Income!$B$4:$B${3+N},Income!$B$4:$B${3+N}&""))'
for rr in (5,6,7):
    db.cell(rr,4).fill=PatternFill("solid",fgColor=LIGHT); db.cell(rr,4).border=box
    db.cell(rr,4).alignment=Alignment(vertical="center",indent=1)
    db.cell(rr,5).font=Font(F,size=12,bold=True); db.cell(rr,5).border=box
    db.cell(rr,5).alignment=Alignment(horizontal="right",indent=1)
db.cell(6,5).number_format=CUR
db.column_dimensions["D"].width=22; db.column_dimensions["E"].width=14

db["A15"]="Expenses by category"; db["A15"].font=Font(F,size=12,bold=True,color=NAVY)
for i in range(11):
    r=16+i
    db.cell(r,1,f'=IF(Settings!C{11+i}="","",Settings!C{11+i})').font=Font(F,size=10)
    db.cell(r,2,f'=IF($A{r}="","",SUMIF(Expenses!$D$4:$D${3+N},$A{r},Expenses!$E$4:$E${3+N}))')
    db.cell(r,2).number_format=CUR0; db.cell(r,2).font=Font(F,size=10)
    db.cell(r,1).border=box; db.cell(r,2).border=box
db.conditional_formatting.add("B16:B26",DataBarRule(start_type="min",end_type="max",color=BLUE))

ch2=BarChart(); ch2.type="bar"; ch2.title="Expenses by category"; ch2.height=9; ch2.width=14
ch2.y_axis.numFmt=CUR0
d2=Reference(db,min_col=2,min_row=16,max_row=26)
c2=Reference(db,min_col=1,min_row=16,max_row=26)
ch2.add_data(d2); ch2.set_categories(c2); ch2.legend=None
db.add_chart(ch2,"D15")

wb.move_sheet("Dashboard",offset=-1)
wb.save("Freelancer-Finance-Tracker.xlsx")
print("fogli calcolati aggiunti:", wb.sheetnames)
